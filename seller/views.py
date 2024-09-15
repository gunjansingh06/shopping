from django.shortcuts import render, redirect, get_object_or_404
from myadmin.models import *
from customer.models import *
from seller.models import *
from django.contrib.auth.models import User
from django.contrib import auth,messages
from datetime import date
from django.core.files.storage import FileSystemStorage
from django.conf import settings
import os
import datetime
from openpyxl import Workbook
import io
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import FileResponse
#pdf
from django.http import FileResponse,HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch 
from reportlab.lib.pagesizes import letter,A3,A2
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Frame, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import Paragraph


def layout(request):
	context = {}
	return render(request ,'seller/common/layout.html',context)

def dashboard(request):
    context = {}
    return render(request ,'seller/dashboard.html',context)

def saler_reg(request):
    state = State.objects.all()
    city = City.objects.all()
    area = Area.objects.all()
    user = User.objects.all()

    context = {'state':state,'city':city,'area':area,'user':user}
    return render(request,'seller/saler_reg.html',context)

def check_username(request,uname):
    if User.objects.filter(username=uname).exists():
        return False 
    else:
        return True    

def check_email(request,email):
    if User.objects.filter(email=email).exists():
        return False
    else:
        return True


def store_reg(request):
    myfirst_name = request.POST['fname']
    mylast_name = request.POST['lname']
    myemail = request.POST['email']
    myusername = request.POST['username']
    mypassword = request.POST['password']
    mycpassword = request.POST['cpassword']
    
    
    myshop_name= request.POST['shop_name']
    myshop_address= request.POST['shop_address']
    myshop_contact= request.POST['shop_contact']
    mysid= request.POST['state']
    mycid= request.POST['city']
    myaid= request.POST['area']
    myshop_image = request.FILES['shop_image']
    mydate = date.today()
    

    mylocation = os.path.join(settings.MEDIA_ROOT, 'upload')
    obj = FileSystemStorage(location=mylocation)
    obj.save(myshop_image.name, myshop_image)




    # if mypassword==mycpassword:
    #     user = User.objects.create_user(first_name=myfirst_name,last_name=mylast_name,email=myemail,username=myusername,password=mypassword)
    #     Saler_reg.objects.create(shop_name=myshop_name,shop_address=myshop_address,shop_contact=myshop_contact,user_id=user.id,state_id=mysid,city_id=mycid,area_id=myaid,image=myshop_image)
    #     return redirect('/seller/layout')

    # else:
    #     messages.success(request,'invalid username or pass')
    #     return redirect('/seller/saler_reg') 

    if mypassword == mycpassword and check_username(request,myusername) and check_email(request,myemail):
        user = User.objects.create_user(first_name=myfirst_name,last_name=mylast_name,email=myemail,username=myusername,password=mypassword)
        Saler_reg.objects.create(shop_name=myshop_name,shop_address=myshop_address,shop_contact=myshop_contact,user_id=user.id,state_id=mysid,city_id=mycid,area_id=myaid,image=myshop_image,date=mydate)

        return redirect('/seller/login')
    else:
        if check_username(request,myusername) is False:
            messages.error(request,'UserName Already Exists...') 
            return redirect('/seller/saler_reg')
        elif check_email(request,myemail) is False:
            messages.error(request,'Email Already Exists...') 
            return redirect('/seller/saler_reg')
  

def delete_reg(request,id):
    result = Saler_reg.objects.get(pk=id)
    result.delete()
    return redirect('/myadmin/all_salers')

def edit_reg(request,id):
    result = Saler_reg.objects.get(pk=id)
    state = State.objects.all()
    city = City.objects.all()
    area = Area.objects.all()

    context={'result':result,'state':state,'city':city,'area':area}
    return render(request,'seller/edit_reg.html',context)


def update_reg(request,id):
    myfirst_name = request.POST['fname']
    mylast_name = request.POST['lname']
    myemail = request.POST['email']
    myusername = request.POST['username']
   
    
    myshop_name= request.POST['shop_name']
    myshop_address= request.POST['shop_address']
    myshop_contact= request.POST['shop_contact']
    mysid= request.POST['state']
    mycid= request.POST['city']
    myaid= request.POST['area']
    myshop_image = request.FILES['shop_image']

    mylocation = os.path.join(settings.MEDIA_ROOT, 'upload')
    obj = FileSystemStorage(location=mylocation)
    obj.save(myshop_image.name, myshop_image)

    data = {
            'first_name':myfirst_name,
            'last_name':mylast_name,
            'email':myemail,
            'username':myusername
           

    }

    data1 = {
            'shop_name':myshop_name,
            'shop_address':myshop_address,
            'shop_contact':myshop_contact,
            'state_id':mysid,
            'city_id':mycid,
            'area_id':myaid,
            'image':myshop_image


    }
     
    user_id = request.user.id
    saler = Saler_reg.objects.get(user_id=user_id)

    User.objects.update_or_create(pk=user_id,defaults=data)
    Saler_reg.objects.update_or_create(pk=saler.id,defaults=data1)
  

    return redirect('/myadmin/all_salers')

def add_product(request):
    categories = Categories.objects.all()
    subcategories = Subcategories.objects.all()
    context = {'categories':categories,'subcategories':subcategories}
    return render(request,'seller/add_product.html',context)

def store_product(request):
    mycid = request.POST['cid']
    myscid = request.POST['scid']
    mypname = request.POST['pname']
    myimage = request.FILES['image']
    myprice = request.POST['price']
    mydetails = request.POST['details']
    myquantity = request.POST['quantity']
    mysize = request.POST.getlist('size[]')
    mysize1 =",".join(mysize)


    user_id = request.user.id
    saler = Saler_reg.objects.get(user_id=user_id)


    mylocation = os.path.join(settings.MEDIA_ROOT, 'upload')
    obj = FileSystemStorage(location=mylocation)
    obj.save(myimage.name, myimage)

    Product.objects.create(categories_id=mycid,subcategories_id=myscid,product_name=mypname,image=myimage,price=myprice,details=mydetails,quantity=myquantity,saler_reg_id=saler.id,size=mysize1)
    return redirect('/seller/add_product')   


def all_product(request):

    user_id = request.user.id
    saler_id = Saler_reg.objects.get(user_id=user_id)

    
    categories = Categories.objects.all()
    subcategories = Subcategories.objects.all()
    product = Product.objects.filter(saler_reg_id=saler_id)
    context = {'product':product,'categories':categories,'subcategories':subcategories}
    return render(request,'seller/all_product.html',context)

def delete_product(request,id):
    result = Product.objects.get(pk=id)
    result.delete()
    return redirect('/seller/all_product')

def edit_product(request,id):
    result = Product.objects.get(pk=id)
    categories = Categories.objects.all()
    subcategories = Subcategories.objects.all()
    result.size = result.size.split(",")
    

    context={'result':result,'categories':categories,'subcategories':subcategories}
    return render(request,'seller/edit_product.html',context)

def update_product(request,id):
    mycid = request.POST['cid']
    myscid = request.POST['scid']
    mypname = request.POST['pname']
    myimage = request.FILES['image']
    myprice = request.POST['price']
    mydetails = request.POST['details']
    myquantity = request.POST['quantity']

    mysize = request.POST.getlist('size[]')
    mysize1 =",".join(mysize)

    mylocation = os.path.join(settings.MEDIA_ROOT, 'upload')
    obj = FileSystemStorage(location=mylocation)
    obj.save(myimage.name, myimage)
    data={
            'categories_id':mycid,
            'subcategories_id':myscid,
            'product_name':mypname,
            'image':myimage,
            'price':myprice,
            'details':mydetails,
            'quantity':myquantity,
        'size' : mysize1,


    }

    Product.objects.update_or_create(pk=id,defaults=data)
    return redirect('/seller/all_product')

def login(request):
    context = {}
    return render(request,'seller/login.html',context)

def login_check(request):
    if request.method == 'POST':
        myusername = request.POST.get('username')
        mypassword = request.POST.get('password')

        if not myusername or not mypassword:
            messages.error(request, "Username and password are required.")
            return redirect('/seller/login')

        result = auth.authenticate(username=myusername, password=mypassword)
        if result is None:
            messages.error(request, "Invalid Username or Password 🤭")
            return redirect('/seller/login')
        else:
            # Check user's status
            try:
                user_register = Saler_reg.objects.get(user=result)
                user_status = user_register.status  # Assuming user profile has a status field
                if user_status == 'Pending':
                    messages.error(request, "Your account is still pending approval.")
                    return redirect('/seller/login')
                else:
                    auth.login(request, result)
                    return redirect('/seller/dashboard')
            except Saler_reg.DoesNotExist:
                messages.error(request, "User profile does not exist.")
                return redirect('/seller/login')
    else:
        return render(request, 'seller/login.html')

def logout(request):
    auth.logout(request)
    return redirect('/seller/login')

def orders(request):
    user_id = request.user.id
    saler_id = Saler_reg.objects.get(user_id=user_id)
    order = Order.objects.filter(saler_id=saler_id)
    context = {'order':order}
    return render(request,'seller/orders.html',context)

def orders_details(request,id):
    order = get_object_or_404(Order, pk=id)
    order_details = Order_details.objects.filter(order=order)
    billing = get_object_or_404(Billing, order=order)
    shipping = get_object_or_404(Shipping, order=order)

    context = {
        'order': order,
        'order_details': order_details,
        'billing': billing,
        'shipping':shipping,
        }
    return render(request,'seller/orders_details.html',context)

def search_order(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    result = []
    
    if start_date and end_date:
        try:
            # Convert strings to datetime objects
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
            
            # Ensure the end_date includes the entire day by adding one day
            end_date += datetime.timedelta(days=1)

            # Filter Notice objects based on the 'date' field
            order = Order.objects.filter(date__range=[start_date, end_date])
        except ValueError as e:
            print(f"Date conversion error: {e}")
    else:
        print("Start date or end date not provided")

    print(f"Start Date: {start_date}, End Date: {end_date}, order: {order}")

    return render(request, 'seller/orders.html', {'order': order, 'start_date': start_date, 'end_date': end_date})

def generate_excel(request):
    # Get the start and end dates from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    result = []
    
    if start_date and end_date:
        try:
            # Convert strings to datetime objects
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
            
            # Ensure the end_date includes the entire day by adding one day
            end_date += datetime.timedelta(days=1)

            # Filter Order objects based on the 'date' field
            result = Order.objects.filter(date__range=[start_date, end_date])
        except ValueError as e:
            print(f"Date conversion error: {e}")
    else:
        print("Start date or end date not provided")
    
    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    
    # Add headers
    headers = ['ID', 'Date', 'Amount', 'Status', 'Payment Method', 'Saler Name', 'User name']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data to the sheet
    for row_num, order in enumerate(result, 2):
        ws.cell(row=row_num, column=1, value=order.id)
        ws.cell(row=row_num, column=2, value=order.date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=3, value=order.amount)
        ws.cell(row=row_num, column=4, value=order.status)
        ws.cell(row=row_num, column=5, value=order.pay_method)
        ws.cell(row=row_num, column=6, value=order.saler_id)
        ws.cell(row=row_num, column=7, value=order.user_id)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the Excel file to a buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    return FileResponse(buf, as_attachment=True, filename='orders.xlsx')

def generate_pdf_orders(request):
    # Get the start and end dates from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    result = []
    
    if start_date and end_date:
        try:
            # Convert strings to datetime objects
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
            
            # Ensure the end_date includes the entire day by adding one day
            end_date += datetime.timedelta(days=1)

            # Filter Order objects based on the 'date' field
            result = Order.objects.filter(date__range=[start_date, end_date])
        except ValueError as e:
            print(f"Date conversion error: {e}")
    else:
        print("Start date or end date not provided")
    
    # Create a PDF buffer
    buf = BytesIO()
    
    # Create a SimpleDocTemplate for the PDF
    doc = SimpleDocTemplate(buf, pagesize=A3)
    
    # Create a list to hold the PDF elements
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    header_style = styles['Heading1']
    normal_style = styles['Normal']
    
    # Add a title
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
    'HeaderStyle',
    parent=styles['Title'],
    alignment=TA_CENTER,
    fontName='Helvetica-Bold',
    fontSize=15
)
    elements.append(Paragraph('Order List', header_style))
    
    # Define table data
    headers = ['ID', 'Date', 'Amount', 'Status', 'Payment Method', 'Saler Name', 'User Name']
    data = [headers]
    
    for order in result:
        row = [
            order.id,
            order.date.strftime('%Y-%m-%d'),
            order.amount,
            order.status,
            order.pay_method,
            order.saler.user.first_name,
            order.user.first_name,
        ]
        data.append(row)
    
    # Create a table with the data
    table = Table(data)
    
    # Define the table style
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f2dfdb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(table_style)
    
    # Add the table to the elements
    elements.append(table)
    
    # Build the PDF
    doc.build(elements)
    
    # Set the buffer position to the beginning
    buf.seek(0)
    
    # Return the PDF as a response
    return HttpResponse(buf, content_type='application/pdf', headers={'Content-Disposition': 'attachment; filename="orders.pdf"'})